import asyncio
from pyppeteer import launch
from openpyxl import Workbook

wb=Workbook()#创建一个workbook
# 通过openpyxl模块创建workbook时，无需本地事先创建好excel，它会直接创建一个新的excel文件
# 创建workbook时，会至少包含一个worksheet
# 注意：openpyxl模块只支持.xlsx,.xlsm,.xltx,.xltm格式
# ws1=wb.create_sheet()#当前workbook结尾处追加新的sheet,名称是自动赋值
# ws2=wb.create_sheet(0)
# ws1.title="new title"#指定sheet名字
# ws2.title = "new title"#指定sheet名称

# ws3=wb.create_sheet("MyNewSheet1")#创建直接赋值名字
# ws4 = wb.create_sheet("MyNewSheet2", 0)
# ws.sheet_properties.tabColor = "1072BA"  #设定worksheet的颜色，设定值为RRGGBB颜色代码
# target = wb.copy_worksheet(source) #拷贝当前worksheet

sheet=wb.active#获取当前workbook的第一个worksheet,默认的索引值0它可以改变

sheet.title = '中国大学排名'
sheet['A1'].value='排名'
sheet['B1'].value='学校名称'
sheet['C1'].value = '省市'
sheet['D1'].value = '学校类型'
sheet['E1'].value = '总分'
sheet['F1'].value = '办学层次'
async def main():
	id=1
	browser=await launch({'headless':False,'args':['--disable-infobars','--window-size=1024,800']})
	page=await browser.newPage()
	await page.setViewport({'width': 1024, 'height': 800})
	await page.goto('https://www.shanghairanking.cn/rankings/bcur/2020')
	#xpath获取表格的位置
	tbody=await page.xpath(r'//*[@id="content-box"]/div[2]/table/tbody//tr')
	for i in tbody:
		#获取文本
		# 获取表格文本        
        # 获取文本：方法一，通过getProperty方法获取
        # title_str1 = await (await item.getProperty('textContent')).jsonValue()
        # 获取文本：方法二，通过evaluate方法获取
        # title_str2 = await page.evaluate('item => item.textContent', item)
        # 获取链接：通过getProperty方法获取
        # title_link = await (await item.getProperty('href')).jsonValue()
		title_str1=await (await i.getProperty('textContent')).jsonValue()#获取到了所有文本，需要切割一下
		srt=title_str1.splitlines()#分割，- 在输出结果里是否去掉行界符('\r', '\r\n', \n'等)，默认为 False，不包含行界符，如果为 True，则保留行界符。
		id+=1
		#print(srt)
		sheet['A%s' % (id)].value=id-1
		sheet['B%s' % (id)].value = srt[2].strip()
		sheet['C%s' % (id)].value = srt[3].strip()
		sheet['D%s' % (id)].value = srt[5].strip()
		sheet['E%s' % (id)].value = srt[7].strip()
		sheet['F%s' % (id)].value = srt[9].strip()

	await browser.close()
asyncio.get_event_loop().run_until_complete(main())
wb.save((r'D:\新建文件夹\pythontext\data\中国大学排名.xlsx'))


